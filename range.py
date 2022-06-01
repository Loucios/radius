from openpyxl import load_workbook
from classes import Connection


def main():
    wb = load_workbook(filename='RET.xlsx', data_only=True)
    # ws = wb['1']
    rng = wb.defined_names.get('Table1', scope=wb.sheetnames.index('1'))
    rng_dict = dict(rng.destinations)
    dest = rng_dict['1']
    private_range = wb['1'][dest]

    tbl_data = []
    for row in private_range:
        cell_values = []
        for cell in row:
            cell_values.append(cell.value)
        tbl_data.append(Connection(*cell_values))

    print(tbl_data)
    return tbl_data


if __name__ == '__main__':
    main()
